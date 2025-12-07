"""

Full-AI Statement of Values (SOV) parsing engine.

- Loads Excel workbooks
- Sends sheet data to an AI agent
- Receives structured JSON for:
    - Property-level summary
    - Building-level records
- Wraps results in SOVParser.parse_excel()
"""

import os
from dataclasses import dataclass
from typing import Optional, List, Dict, Any, Tuple

from openpyxl import load_workbook

from app.parser.mapping_agent_single import call_sov_sheet_agent
import json
import re

# =====================================================================
# 1. Canonical Fields
# =====================================================================


BUILDING_FIELDS = [
    "building_number",
    "location_full_address",
    "location_address",
    "location_city",
    "location_state",
    "location_zip",
    "units_per_building",
    "replacement_cost_tiv",
    "num_units",
    "livable_sq_ft",
    "garage_sq_ft",
]

NUMERIC_BUILDING_FIELDS = {
    "units_per_building",
    "replacement_cost_tiv",
    "num_units",
    "livable_sq_ft",
    "garage_sq_ft",
}

# =====================================================================
# 2. Dataclasses
# =====================================================================


@dataclass
class BuildingRecord:
    building_number: Optional[str] = None
    location_full_address: Optional[str] = None
    location_address: Optional[str] = None
    location_city: Optional[str] = None
    location_state: Optional[str] = None
    location_zip: Optional[str] = None
    units_per_building: Optional[float] = None
    replacement_cost_tiv: Optional[float] = None
    num_units: Optional[float] = None
    livable_sq_ft: Optional[float] = None
    garage_sq_ft: Optional[float] = None

    def to_row(self):
        return [getattr(self, f) for f in BUILDING_FIELDS]


# =====================================================================
# 3. Helpers
# =====================================================================


def _rows_from_sheet(ws) -> List[List[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _coerce_float(v: Any) -> Optional[float]:
    """
    Convert values like "1,234", "  123.45 ", or 123 to float.
    Treat empty strings, 'n/a', '-', etc. as None.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().lower()
        if s in {"", "n/a", "na", "none", "null", "-", "--"}:
            return None
        s = s.replace(",", "")
        return float(s)
    except Exception:
        return None


def _building_from_ai_dict(d: Dict[str, Any]) -> BuildingRecord:
    data: Dict[str, Any] = {f: None for f in BUILDING_FIELDS}
    data.update({k: d.get(k) for k in d.keys() if k in BUILDING_FIELDS})

    for f in NUMERIC_BUILDING_FIELDS:
        if f in data:
            data[f] = _coerce_float(data[f])

    return BuildingRecord(**{k: data.get(k) for k in BUILDING_FIELDS})


def normalize_address(s: str):
    """Normalize addresses heavily for fuzzy matching."""
    if s is None:
        return None
    s = str(s)

    # unify whitespace
    s = " ".join(s.split())

    # unify dashes
    s = re.sub(r"[–—−]", "-", s)

    # lowercase
    s = s.lower()

    # remove trailing punctuation
    s = s.rstrip(",.; ")

    # normalize apt/unit/# indicators
    s = s.replace(" apt ", " unit ")
    s = s.replace(" #", " unit ")
    s = s.replace(" suite ", " unit ")

    # collapse duplicates
    s = re.sub(r"unit\s+unit", "unit", s)

    return " ".join(s.split())


def _to_float_if_possible(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except:
        return None


def _normalize_string(v):
    if v is None:
        return None
    s = str(v).lower().strip()
    s = " ".join(s.split())
    s = re.sub(r"[^\w\s\-]", "", s)  # remove punctuation except dash
    return s


def fuzzy_numeric_equal(a, b, tol=0.5):
    fa = _to_float_if_possible(a)
    fb = _to_float_if_possible(b)
    if fa is None or fb is None:
        return False

    # auto-round large values (sqft, TIV)
    if abs(fa) > 1000:
        fa = round(fa, 1)
        fb = round(fb, 1)

    return abs(fa - fb) <= tol


def _fuzzy_equal(a, b):
    """Smarter fuzzy matching including partial address cleaning."""

    # Both None → equal
    if a is None and b is None:
        return True

    # One None → not equal (unless numeric override handles it)
    if a is None or b is None:
        return False

    # ---- NUMERIC MATCH FIRST ----
    if fuzzy_numeric_equal(a, b):
        return True

    # ---- ADDRESS NORMALIZATION ----
    na = normalize_address(a)
    nb = normalize_address(b)

    # Exact normalized match
    if na == nb:
        return True

    # NEW: Handle noisy addresses like "STREET, null, null"
    cleaned_nb = nb.replace("null", "").replace(",", " ")
    cleaned_nb = " ".join(cleaned_nb.split())  # collapse spaces

    cleaned_na = na.replace("null", "").replace(",", " ")
    cleaned_na = " ".join(cleaned_na.split())

    # If either cleaned form contains the other → match
    if cleaned_na in cleaned_nb or cleaned_nb in cleaned_na:
        return True

    # ---- GENERAL STRING MATCH ----
    sa = _normalize_string(a)
    sb = _normalize_string(b)

    # NEW: Apply same cleaning to general strings
    cleaned_sa = sa.replace("null", "").replace(",", "")
    cleaned_sb = sb.replace("null", "").replace(",", "")

    cleaned_sa = " ".join(cleaned_sa.split())
    cleaned_sb = " ".join(cleaned_sb.split())

    if cleaned_sa == cleaned_sb:
        return True

    return False


def compare_model_outputs(
    result_a: Dict[str, Any], result_b: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Compare two model outputs using fuzzy matching rules.
    Only keep fields that match (after normalization).
    """
    buildings_a = result_a.get("buildings", [])
    buildings_b = result_b.get("buildings", [])

    final = {"buildings": []}

    max_len = max(len(buildings_a), len(buildings_b))

    for idx in range(max_len):
        rec_a = buildings_a[idx] if idx < len(buildings_a) else {}
        rec_b = buildings_b[idx] if idx < len(buildings_b) else {}

        matching = {}
        mismatches = {}

        for key in BUILDING_FIELDS:
            v1 = rec_a.get(key)
            v2 = rec_b.get(key)

            # --- CASE 1: both missing → HARD MISMATCH ---
            if v1 is None and v2 is None:
                mismatches[key] = {"gpt_5_1": None, "gpt_4o_mini": None}
                matching[key] = None
                continue

            # --- CASE 2: fuzzy match → OK ---
            if _fuzzy_equal(v1, v2):
                matching[key] = v1
                continue

            # --- CASE 3: ONE model missing → accept the non-null value (NO ALERT) ---
            if v1 is None and v2 is not None:
                matching[key] = v2
                continue

            if v2 is None and v1 is not None:
                matching[key] = v1
                continue

            # --- CASE 4: BOTH have values but DIFFER → ALERT ---
            mismatches[key] = {"gpt_5_1": v1, "gpt_4o_mini": v2}

        if mismatches:
            print(f"\n⚠️ LLM MISMATCH DETECTED (sheet building index {idx}):")
            print(json.dumps(mismatches, indent=4))

        final["buildings"].append(matching)

    return final


# =====================================================================
# 4. General Sheet Detection & Metadata Extraction
# =====================================================================


def extract_property_metadata(ai_result: Dict[str, Any]) -> Dict[str, Any]:
    buildings = ai_result.get("buildings", [])
    if len(buildings) == 1 and not buildings[0].get("building_number"):
        b = buildings[0]
        return {
            k: b.get(k)
            for k in [
                "location_full_address",
                "location_address",
                "location_city",
                "location_state",
                "location_zip",
            ]
            if b.get(k)
        }
    return {}


# =====================================================================
# 5. Workbook Parsing (Full-AI)
# =====================================================================


def parse_workbook(path: str) -> Tuple[List[BuildingRecord]]:
    wb = load_workbook(path, data_only=True)

    building_records: List[BuildingRecord] = []

    source_file = os.path.basename(path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _rows_from_sheet(ws)

        # --- NEW: Run both models ---
        ai_gpt5 = call_sov_sheet_agent(
            source_file=source_file, sheet_name=sheet_name, rows=rows, model="gpt-5.1"
        )

        ai_mini = call_sov_sheet_agent(
            source_file=source_file,
            sheet_name=sheet_name,
            rows=rows,
            model="gpt-4o-mini",
        )

        # Compare them & keep only matching fields
        ai_result = compare_model_outputs(ai_gpt5, ai_mini)

        # Extract final buildings
        sheet_bldgs = ai_result.get("buildings", [])

        for b in sheet_bldgs:
            # Guarantee required fields
            b.setdefault("source_file", source_file)
            b.setdefault("sheet_name", sheet_name)

            building_records.append(_building_from_ai_dict(b))

    return building_records


# =====================================================================
# 5. Public Interface
# =====================================================================


class SOVParser:
    """
    Full-AI SOV parser interface.
    Use:
        parser = SOVParser()
        result = parser.parse_excel("path/to/file.xlsx")
    """

    def parse_excel(self, path: str) -> Dict[str, Any]:
        building_records = parse_workbook(path)
        return {
            "buildings": [b.__dict__ for b in building_records],
        }
