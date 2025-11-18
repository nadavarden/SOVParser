import os
import json
import time
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

# Import your two agents
from app.parser.mapping_agent import call_sov_sheet_agent as call_sov_sheet_agent_multi
from app.parser.mapping_agent_single import call_sov_sheet_agent as call_sov_sheet_agent_single


# ---------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------

TEST_DATA_DIR = "tests/data"
GROUND_TRUTH_DIR = "tests/ground_truth"
RESULTS_DIR = "tests/results"

os.makedirs(RESULTS_DIR, exist_ok=True)

# If you want a specific model for the agents, put it here.
# If your agents ignore this or have their own default, that's fine.
DEFAULT_MODEL = None  # e.g. "gpt-4.1-mini"


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def rows_from_sheet(ws) -> List[List[Any]]:
    """Convert an openpyxl worksheet to a list of rows (values only)."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


def compute_similarity(a: Any, b: Any) -> float:
    """
    Simple similarity:

    - If both are numeric → 1 - relative error
    - Else → case-insensitive string similarity
      (very lightweight; you can swap for rapidfuzz if you like)
    """
    if a is None and b is None:
        return 1.0
    if a is None or b is None:
        return 0.0

    # Try numeric
    try:
        a_num = float(a)
        b_num = float(b)
        denom = abs(a_num) + 1e-6
        return max(0.0, 1.0 - abs(a_num - b_num) / denom)
    except Exception:
        pass

    sa = str(a).strip().lower()
    sb = str(b).strip().lower()
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0

    # Simple token-based similarity
    # (you can replace with difflib if you prefer)
    matches = sum(1 for x, y in zip(sa, sb) if x == y)
    return matches / max(len(sa), len(sb))


def run_agent_on_workbook(
    xlsx_path: str,
    call_fn,
    model: str = None,
) -> Tuple[Dict[str, List[Dict[str, Any]]], float]:
    """
    Run one agent on all sheets of the workbook and combine the results.
    Returns ({'properties': [...], 'buildings': [...]}, runtime_seconds)
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    source_file = os.path.basename(xlsx_path)

    all_properties: List[Dict[str, Any]] = []
    all_buildings: List[Dict[str, Any]] = []

    start = time.time()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = rows_from_sheet(ws)

        # Agent API: call_sov_sheet_agent(source_file, sheet_name, rows, model=None)
        result = call_fn(
            source_file=source_file,
            sheet_name=sheet_name,
            rows=rows,
            model=model,
        )

        sheet_props = result.get("properties", []) or []
        sheet_bldgs = result.get("buildings", []) or []

        all_properties.extend(sheet_props)
        all_buildings.extend(sheet_bldgs)

    runtime = time.time() - start

    return {
        "properties": all_properties,
        "buildings": all_buildings,
    }, runtime


def evaluate_record_list(
    expected_records: List[Dict[str, Any]],
    predicted_records: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Compare a list of expected records to predicted ones, index-aligned.
    Returns accuracy, coverage, and per-record/field details.
    """
    total_fields = 0
    total_similarity = 0.0
    filled_fields = 0

    per_record: List[Dict[str, Dict[str, Any]]] = []

    max_len = max(len(expected_records), len(predicted_records))

    for i in range(max_len):
        gt = expected_records[i] if i < len(expected_records) else {}
        pr = predicted_records[i] if i < len(predicted_records) else {}

        record_detail: Dict[str, Dict[str, Any]] = {}

        # Only evaluate fields that exist in ground truth
        for field, expected_val in gt.items():
            predicted_val = pr.get(field)

            sim = compute_similarity(expected_val, predicted_val)
            total_similarity += sim
            total_fields += 1

            if predicted_val not in (None, "", []):
                filled_fields += 1

            record_detail[field] = {
                "expected": expected_val,
                "predicted": predicted_val,
                "similarity": round(sim, 4),
            }

        per_record.append(record_detail)

    accuracy = (total_similarity / total_fields) if total_fields else 0.0
    coverage = (filled_fields / total_fields) if total_fields else 0.0

    return {
        "num_expected_records": len(expected_records),
        "num_predicted_records": len(predicted_records),
        "total_fields": total_fields,
        "total_similarity": total_similarity,
        "filled_fields": filled_fields,
        "accuracy": round(accuracy, 3),
        "coverage": round(coverage, 3),
        "records": per_record,
    }


def evaluate_file(xlsx_path: str, ground_truth_path: str) -> Dict[str, Any]:
    """
    Evaluate both agents on a single workbook.
    """
    filename = os.path.basename(xlsx_path)
    print(f"\n========== {filename} ==========")

    with open(ground_truth_path, "r") as f:
        gt = json.load(f)

    gt_properties = gt.get("properties", []) or []
    gt_buildings = gt.get("buildings", []) or []

    results: Dict[str, Any] = {}

    agents = {
        "multi_agent": call_sov_sheet_agent_multi,
        "single_agent": call_sov_sheet_agent_single,
    }

    for agent_name, fn in agents.items():
        print(f"\nRunning {agent_name}...")

        try:
            predictions, runtime = run_agent_on_workbook(
                xlsx_path=xlsx_path,
                call_fn=fn,
                model=DEFAULT_MODEL,
            )
        except Exception as e:
            print(f"  ❌ Agent failed: {e}")
            results[agent_name] = {"error": str(e)}
            continue

        pred_props = predictions.get("properties", []) or []
        pred_bldgs = predictions.get("buildings", []) or []

        props_eval = evaluate_record_list(gt_properties, pred_props)
        bldgs_eval = evaluate_record_list(gt_buildings, pred_bldgs)

        # Global accuracy/coverage weighted by number of fields
        total_fields = props_eval["total_fields"] + bldgs_eval["total_fields"]
        total_similarity = props_eval["total_similarity"] + bldgs_eval["total_similarity"]
        filled_fields = props_eval["filled_fields"] + bldgs_eval["filled_fields"]

        overall_accuracy = (total_similarity / total_fields) if total_fields else 0.0
        overall_coverage = (filled_fields / total_fields) if total_fields else 0.0

        print(f"  Runtime:           {runtime:.2f} s")
        print(f"  Overall accuracy:  {overall_accuracy:.3f}")
        print(f"  Overall coverage:  {overall_coverage:.3f}")
        print(f"  Props:  acc={props_eval['accuracy']:.3f}, cov={props_eval['coverage']:.3f}, "
              f"records gt={props_eval['num_expected_records']}, pred={props_eval['num_predicted_records']}")
        print(f"  Bldgs:  acc={bldgs_eval['accuracy']:.3f}, cov={bldgs_eval['coverage']:.3f}, "
              f"records gt={bldgs_eval['num_expected_records']}, pred={bldgs_eval['num_predicted_records']}")

        results[agent_name] = {
            "runtime_seconds": round(runtime, 2),
            "overall_accuracy": round(overall_accuracy, 3),
            "overall_coverage": round(overall_coverage, 3),
            "properties": props_eval,
            "buildings": bldgs_eval,
        }

    # Write detailed JSON
    out_path = os.path.join(
        RESULTS_DIR,
        os.path.splitext(filename)[0] + "_eval.json"
    )
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)

    print(f"\nSaved detailed results to: {out_path}")
    return results


def run_all():
    if not os.path.isdir(TEST_DATA_DIR):
        print(f"Test data dir not found: {TEST_DATA_DIR}")
        return

    files = [
        f for f in os.listdir(TEST_DATA_DIR)
        if f.lower().endswith(".xlsx")
    ]

    if not files:
        print(f"No .xlsx files found in {TEST_DATA_DIR}")
        return

    print(f"Found {len(files)} workbook(s):")
    for f in files:
        print(f"  - {f}")

    for f in files:
        xlsx_path = os.path.join(TEST_DATA_DIR, f)
        gt_path = os.path.join(GROUND_TRUTH_DIR, os.path.splitext(f)[0] + ".json")

        if not os.path.exists(gt_path):
            print(f"\n⚠ No ground truth for {f} ({gt_path} missing) — skipping.")
            continue

        evaluate_file(xlsx_path, gt_path)


if __name__ == "__main__":
    run_all()