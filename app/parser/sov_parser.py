"""
sov_parser.py

Core Statement of Values (SOV) parsing engine.

- Handles multi-sheet Excel files
- Detects building header tables
- Extracts property-level fields
- Extracts building-level rows
"""

import os
import re
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple, Any

from openpyxl import load_workbook

# OPTIONAL – uncomment these if you later want AI mapping enabled
# from app.parser.mapping_agent import call_mapping_agent
# from app.utils import collect_top_left_cells

# =====================================================================
# 1. Canonical Fields
# =====================================================================

PROPERTY_FIELDS = [
    "source_file",
    "sheet_name",
    "number_of_buildings",
    "roof_type",
    "building_valuation_type",
    "building_replacement_cost",
    "blanket_outdoor_property",
    "business_personal_property",
    "total_insurable_value",
    "general_liability",
    "building_ordinance_a",
    "building_ordinance_b",
    "building_ordinance_c",
    "equipment_breakdown",
    "sewer_or_drain_backup",
    "business_income",
    "hired_and_non_owned_auto",
    "playgrounds_number",
    "streets_miles",
    "pools_number",
    "spas_number",
    "wader_pools_number",
    "restroom_building_sq_ft",
    "guardhouse_sq_ft",
    "clubhouse_sq_ft",
    "fitness_center_sq_ft",
    "tennis_courts_number",
    "basketball_courts_number",
    "other_sport_courts_number",
    "walking_biking_trails_miles",
    "lakes_or_ponds_number",
    "boat_docks_and_slips_number",
    "dog_parks_number",
    "elevators_number",
    "commercial_exposure_sq_ft",
]

BUILDING_FIELDS = [
    "source_file",
    "sheet_name",
    "row_index",
    "building_number",
    "location_full_address",
    "location_address",
    "location_city",
    "location_state",
    "location_zip",
    "lat",
    "long",
    "betterview_id",
    "betterview_building_number",
    "units_per_building",
    "replacement_cost_tiv",
    "num_units",
    "livable_sq_ft",
    "garage_sq_ft",
    "commercial_sq_ft",
    "building_class",
    "parking_type",
    "roof_type",
    "smoke_detectors",
    "sprinklered",
    "year_of_construction",
    "number_of_stories",
    "construction_type",
]

# =====================================================================
# 2. Header Normalization
# =====================================================================

def normalize_header(v: str) -> str:
    v = v.strip().lower()
    v = re.sub(r"\s+", " ", v)
    return v.replace(":", "")

# =====================================================================
# 3. Building Header Mapping (Heuristic)
# =====================================================================

# All keys should already be "normalized" (lowercase, no extra spaces / colons)
BUILDING_HEADER_MAP: Dict[str, str] = {
    # building id
    "bldg #": "building_number",
    "bldg": "building_number",
    "building #": "building_number",
    "building number": "building_number",
    "bldg no.": "building_number",

    # address
    "location full address": "location_full_address",
    "address #": "location_address",
    "address": "location_address",
    "address name": "location_address",
    "street": "location_address",

    "city": "location_city",
    "st": "location_state",
    "state": "location_state",
    "zip": "location_zip",
    "zip code": "location_zip",

    # geo
    "lat": "lat",
    "latitude": "lat",
    "long": "long",
    "longitude": "long",

    # betterview
    "betterview id": "betterview_id",
    "betterview building number": "betterview_building_number",

    # units / size
    "units per building": "units_per_building",
    "number of units": "num_units",
    "# of units": "num_units",
    "units": "num_units",

    # values
    "replacement cost": "replacement_cost_tiv",
    "building tiv": "replacement_cost_tiv",
    "tiv": "replacement_cost_tiv",

    # area
    "livable sq ft": "livable_sq_ft",
    "livable sq. ft.": "livable_sq_ft",
    "livable square footage": "livable_sq_ft",
    "garage sq ft": "garage_sq_ft",
    "garage sq. ft.": "garage_sq_ft",
    "garage square footage": "garage_sq_ft",
    "commercial sq ft": "commercial_sq_ft",

    # misc building attrs
    "building class": "building_class",
    "parking type": "parking_type",
    "roof type": "roof_type",
    "smoke detectors": "smoke_detectors",
    "sprinklered": "sprinklered",
    "year of construction": "year_of_construction",
    "year built": "year_of_construction",
    "number of stories": "number_of_stories",
    "# of stories": "number_of_stories",
    "stories": "number_of_stories",
    "construction type": "construction_type",
    "construction": "construction_type",
}

NUMERIC_BUILDING_FIELDS = {
    "lat", "long", "units_per_building", "replacement_cost_tiv",
    "num_units", "livable_sq_ft", "garage_sq_ft", "commercial_sq_ft",
    "year_of_construction", "number_of_stories"
}

# =====================================================================
# 4. Property Label Mapping
# =====================================================================

PROPERTY_CELL_MAP: Dict[str, str] = {
    "number of buildings": "number_of_buildings",
    "# of buildings": "number_of_buildings",

    "roof type": "roof_type",

    "building valuation type": "building_valuation_type",

    # replacement cost / tiv – MANY variants
    "building replacement cost": "building_replacement_cost",
    "replacement cost": "building_replacement_cost",
    "building replacement cost ($)": "building_replacement_cost",
    "bldg replacement cost": "building_replacement_cost",
    "bldg repl cost": "building_replacement_cost",
    "rc": "building_replacement_cost",
    "rc value": "building_replacement_cost",

    "blanket outdoor property": "blanket_outdoor_property",
    "business personal property": "business_personal_property",
    "bpp": "business_personal_property",
    "total insurable value": "total_insurable_value",
    "tiv": "total_insurable_value",

    "general liability": "general_liability",
    "building ordinance a": "building_ordinance_a",
    "building ordinance b": "building_ordinance_b",
    "building ordinance c": "building_ordinance_c",
    "a , b & c limits": "building_ordinance_a",

    "equipment breakdown": "equipment_breakdown",
    "sewer or drain back up": "sewer_or_drain_backup",
    "sewer backup": "sewer_or_drain_backup",

    "business income": "business_income",
    "actual loss sustained total": "business_income",
    "als": "business_income",

    "hired and non-owned auto": "hired_and_non_owned_auto",

    "playgrounds": "playgrounds_number",
    "streets miles": "streets_miles",
    "pools": "pools_number",
    "spas": "spas_number",
    "wader pools": "wader_pools_number",
    "restroom building sq ft": "restroom_building_sq_ft",
    "guardhouse sq ft": "guardhouse_sq_ft",
    "clubhouse sq ft": "clubhouse_sq_ft",
    "fitness center sq ft": "fitness_center_sq_ft",
    "tennis courts": "tennis_courts_number",
    "basketball courts": "basketball_courts_number",
    "other sport courts": "other_sport_courts_number",
    "walking / biking trails": "walking_biking_trails_miles",
    "walking/biking trails": "walking_biking_trails_miles",
    "lakes or ponds": "lakes_or_ponds_number",
    "boat docks and slips": "boat_docks_and_slips_number",
    "dog parks": "dog_parks_number",
    "elevators": "elevators_number",
    "commercial exposure sq ft": "commercial_exposure_sq_ft",
}

NUMERIC_PROPERTY_FIELDS = {
    "number_of_buildings", "building_replacement_cost", "blanket_outdoor_property",
    "business_personal_property", "total_insurable_value", "general_liility",
    "building_ordinance_a", "building_ordinance_b", "building_ordinance_c",
    "equipment_breakdown", "sewer_or_drain_backup", "business_income",
    "hired_and_non_owned_auto", "playgrounds_number", "streets_miles",
    "pools_number", "spas_number", "wader_pools_number",
    "restroom_building_sq_ft", "guardhouse_sq_ft", "clubhouse_sq_ft",
    "fitness_center_sq_ft", "tennis_courts_number", "basketball_courts_number",
    "other_sport_courts_number", "walking_biking_trails_miles",
    "lakes_or_ponds_number", "boat_docks_and_slips_number",
    "dog_parks_number", "elevators_number", "commercial_exposure_sq_ft"
}

# =====================================================================
# 5. Dataclasses
# =====================================================================

@dataclass
class PropertyRecord:
    source_file: str
    sheet_name: str
    number_of_buildings: Optional[float] = None
    roof_type: Optional[str] = None
    building_valuation_type: Optional[str] = None
    building_replacement_cost: Optional[float] = None
    blanket_outdoor_property: Optional[float] = None
    business_personal_property: Optional[float] = None
    total_insurable_value: Optional[float] = None
    general_liability: Optional[float] = None
    building_ordinance_a: Optional[float] = None
    building_ordinance_b: Optional[float] = None
    building_ordinance_c: Optional[float] = None
    equipment_breakdown: Optional[float] = None
    sewer_or_drain_backup: Optional[float] = None
    business_income: Optional[float] = None
    hired_and_non_owned_auto: Optional[float] = None
    playgrounds_number: Optional[float] = None
    streets_miles: Optional[float] = None
    pools_number: Optional[float] = None
    spas_number: Optional[float] = None
    wader_pools_number: Optional[float] = None
    restroom_building_sq_ft: Optional[float] = None
    guardhouse_sq_ft: Optional[float] = None
    clubhouse_sq_ft: Optional[float] = None
    fitness_center_sq_ft: Optional[float] = None
    tennis_courts_number: Optional[float] = None
    basketball_courts_number: Optional[float] = None
    other_sport_courts_number: Optional[float] = None
    walking_biking_trails_miles: Optional[float] = None
    lakes_or_ponds_number: Optional[float] = None
    boat_docks_and_slips_number: Optional[float] = None
    dog_parks_number: Optional[float] = None
    elevators_number: Optional[float] = None
    commercial_exposure_sq_ft: Optional[float] = None

    def to_row(self):
        return [getattr(self, f) for f in PROPERTY_FIELDS]


@dataclass
class BuildingRecord:
    source_file: str
    sheet_name: str
    row_index: int
    building_number: Optional[str] = None
    location_full_address: Optional[str] = None
    location_address: Optional[str] = None
    location_city: Optional[str] = None
    location_state: Optional[str] = None
    location_zip: Optional[str] = None
    lat: Optional[float] = None
    long: Optional[float] = None
    betterview_id: Optional[str] = None
    betterview_building_number: Optional[str] = None
    units_per_building: Optional[float] = None
    replacement_cost_tiv: Optional[float] = None
    num_units: Optional[float] = None
    livable_sq_ft: Optional[float] = None
    garage_sq_ft: Optional[float] = None
    commercial_sq_ft: Optional[float] = None
    building_class: Optional[str] = None
    parking_type: Optional[str] = None
    roof_type: Optional[str] = None
    smoke_detectors: Optional[str] = None
    sprinklered: Optional[str] = None
    year_of_construction: Optional[int] = None
    number_of_stories: Optional[float] = None
    construction_type: Optional[str] = None

    def to_row(self):
        return [getattr(self, f) for f in BUILDING_FIELDS]

# =====================================================================
# 6. Helper Functions
# =====================================================================

def parse_numeric(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except Exception:
        return None


def is_totals_row(row) -> bool:
    """
    Heuristic: detect footer / totals / misc rows that are not actual building rows.
    """
    text = " ".join(str(c).lower() for c in row if isinstance(c, str))
    return any(t in text for t in ("total", "misc", "statement of values"))

# =====================================================================
# 7. Detect Header Row
# =====================================================================

def detect_header_row(ws) -> Tuple[Optional[int], List[List[Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    best_idx: Optional[int] = None
    best_score = -1

    for i, row in enumerate(rows):
        score = 0
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                n = normalize_header(cell)
                if n in BUILDING_HEADER_MAP:
                    score += 2
                elif any(k in n for k in ("bldg", "street", "address", "zip", "city", "units", "livable")):
                    score += 1
        if score > best_score:
            best_score = score
            best_idx = i

    # require at least a minimal score, otherwise treat as "no header"
    if best_idx is None or best_score < 2:
        return None, rows

    return best_idx, rows

# =====================================================================
# 8. Column Mapping
# =====================================================================

def build_column_mapping(header_row) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if not isinstance(cell, str) or not cell.strip():
            continue
        n = normalize_header(cell)

        if n in BUILDING_HEADER_MAP:
            mapping[idx] = BUILDING_HEADER_MAP[n]
            continue

        # safe substring fallback for a few known keys
        for k, field in BUILDING_HEADER_MAP.items():
            if k in n and len(k) > 3:
                mapping[idx] = field
                break

    return mapping

# =====================================================================
# 9. Property Extraction
# =====================================================================

def _set_property_value(values: Dict[str, Any], target: str, raw_value: Any):
    if raw_value is None:
        return
    if target in NUMERIC_PROPERTY_FIELDS:
        parsed = parse_numeric(raw_value)
        if parsed is not None:
            values[target] = parsed
    else:
        s = str(raw_value).strip()
        if s:
            values[target] = s


def extract_property_info(path: str, ws, num_buildings: int) -> PropertyRecord:
    """
    Scan the sheet for property-level fields using a combination of:
    - label cell + neighbor value (right / below / diagonal)
    - numeric cell + label above (common in SOV summary blocks)
    """
    values: Dict[str, Any] = {f: None for f in PROPERTY_FIELDS}
    values["source_file"] = os.path.basename(path)
    values["sheet_name"] = ws.title
    values["number_of_buildings"] = num_buildings

    max_r = min(ws.max_row, 300)   # safety cap
    max_c = min(ws.max_column, 25)

    # 1) Label-driven extraction: "Field:" and value nearby
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str) or not v.strip():
                continue

            norm = normalize_header(v)
            target = None

            if norm in PROPERTY_CELL_MAP:
                target = PROPERTY_CELL_MAP[norm]
            else:
                for k, f in PROPERTY_CELL_MAP.items():
                    if k in norm:
                        target = f
                        break

            if not target:
                continue

            # collect neighbor candidates:
            neighbors: List[Any] = []

            # right 1..3 cells
            for dc in range(1, 4):
                if c + dc <= max_c:
                    neighbors.append(ws.cell(row=r, column=c + dc).value)

            # down 1..3 cells
            for dr in range(1, 4):
                if r + dr <= max_r:
                    neighbors.append(ws.cell(row=r + dr, column=c).value)

            # diagonals (down-right)
            for dr in range(1, 3):
                for dc in range(1, 3):
                    if r + dr <= max_r and c + dc <= max_c:
                        neighbors.append(ws.cell(row=r + dr, column=c + dc).value)

            val = next((x for x in neighbors if x not in (None, "")), None)
            if val is None:
                continue

            # only overwrite if not already set (first match wins)
            if values.get(target) is None:
                _set_property_value(values, target, val)

    # 2) Numeric-driven extraction: value cell with label above it
    for r in range(2, max_r + 1):  # start at row 2 because we look at r-1
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                continue

            is_numeric_like = isinstance(v, (int, float))
            if not is_numeric_like:
                # allow numeric-looking strings
                if parse_numeric(v) is not None:
                    is_numeric_like = True

            if not is_numeric_like:
                continue

            label_above = ws.cell(row=r - 1, column=c).value
            if not isinstance(label_above, str) or not label_above.strip():
                continue

            norm_above = normalize_header(label_above)
            target = None

            if norm_above in PROPERTY_CELL_MAP:
                target = PROPERTY_CELL_MAP[norm_above]
            else:
                for k, f in PROPERTY_CELL_MAP.items():
                    if k in norm_above:
                        target = f
                        break

            if not target:
                continue

            if values.get(target) is None:
                _set_property_value(values, target, v)

    return PropertyRecord(**values)

# =====================================================================
# 10. Building Parsing
# =====================================================================

def parse_buildings_from_sheet(path: str, ws) -> List[BuildingRecord]:
    records: List[BuildingRecord] = []

    header_idx, all_rows = detect_header_row(ws)
    if header_idx is None:
        return records

    header_row = all_rows[header_idx]
    mapping = build_column_mapping(header_row)

    # If header is poor, optional LLM fallback – left as a hook
    # if len(mapping) < 2:
    #     top_left = collect_top_left_cells(ws)
    #     ai = call_mapping_agent(header_row, all_rows[header_idx+1:header_idx+6], top_left)
    #     if "building_header_mapping" in ai:
    #         mapping = {i: ai["building_header_mapping"].get(str(cell), None)
    #                    for i, cell in enumerate(header_row)}

    for ridx in range(header_idx + 1, len(all_rows)):
        row = all_rows[ridx]

        # skip completely empty rows
        if not any(c not in (None, "") for c in row):
            continue
        # skip total/footer rows
        if is_totals_row(row):
            continue

        rec: Dict[str, Any] = {f: None for f in BUILDING_FIELDS}
        rec["source_file"] = os.path.basename(path)
        rec["sheet_name"] = ws.title
        rec["row_index"] = ridx + 1  # 1-based for Excel

        for col_idx, field in mapping.items():
            if field is None:
                continue
            if col_idx >= len(row):
                continue

            cell = row[col_idx]
            if cell in (None, ""):
                continue

            if field in NUMERIC_BUILDING_FIELDS:
                rec[field] = parse_numeric(cell)
            else:
                rec[field] = str(cell).strip()

        # -----------------------------------------------------------------
        # Row inclusion logic: filter out amenity / non-building rows
        # -----------------------------------------------------------------
        bn = rec.get("building_number")
        addr = rec.get("location_address")
        livable = rec.get("livable_sq_ft")
        units = rec.get("num_units")

        is_valid_bn = False
        if bn is not None:
            bn_str = str(bn).strip()
            # "1", "2", "10"
            if bn_str.isdigit():
                is_valid_bn = True
            # "1A", "2B" pattern
            elif re.match(r"^\d+[A-Za-z]$", bn_str):
                is_valid_bn = True

        # Accept row if it looks like a real building row:
        # - has a valid building_number, and
        # - has at least one of address / livable_sq_ft / units
        if is_valid_bn and (addr or livable or units):
            records.append(BuildingRecord(**rec))

    return records

# =====================================================================
# 11. Workbook Parsing
# =====================================================================

def parse_workbook(path: str):
    wb = load_workbook(path, data_only=True)

    property_records: List[PropertyRecord] = []
    building_records: List[BuildingRecord] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        bldgs = parse_buildings_from_sheet(path, ws)

        if bldgs:
            building_records.extend(bldgs)

            # derive number_of_buildings from unique building_number
            unique_bldg_ids = {
                b.building_number or f"{b.location_address}_{b.location_city}"
                for b in bldgs
            }
            num_buildings = len(unique_bldg_ids)

            prop = extract_property_info(path, ws, num_buildings)
            property_records.append(prop)

    return property_records, building_records

# =====================================================================
# 12. Parser Class (Public Interface)
# =====================================================================

class SOVParser:
    """
    Public parser interface for use in API, CLI, or tests.
    Wraps the workbook-level functions into a clean .parse_excel() method.
    """

    def parse_excel(self, path: str) -> dict:
        property_records, building_records = parse_workbook(path)
        return {
            "properties": [p.__dict__ for p in property_records],
            "buildings": [b.__dict__ for b in building_records],
        }